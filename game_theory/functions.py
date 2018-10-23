"""
Filename: functions.py

Provides key functions to setting up a game between senators.
"""

import csv
import Cayley as cy
import numpy as np
import random
import math
import openpyxl
import xlsxwriter

#For personal reference when using the CSV
nname_of_data_from_csv = ['rank_from_low', 'rank_from_high',
                         'percentile', 'ideology', 'id',
                         'bioguide_id', 'state', 'district', 'name']

__author__ = "\n".join(['Justin Pusztay (pusztayj20@mail.wlu.edu)'])

__all__ = ['make_senate','payoff_matrix1','payoff_matrix2','payoff_matrix3',
           'payoff_matrix4','random_strat_start','game','timestep','data_export',
           'strat_real_imagined_setup','export_data','random_agent']

def make_senate(network,csv_name = 'senatedata.csv'):
    """
    Takes a network and adds each senator to it. It puts their last name
    as the name of the node. Adds ideological score and state as features.

    Notes
    -----
    -> An empty graph network object is best recommended for this function. Since
       it is empty, this function will populate it with senators as nodes.
    -> The idelogical score for all senators are from
       https://www.govtrack.us/congress/members/report-cards/2017/senate/ideology
       These scores are added as a feature to each seantor node.
    -> The state the senator is from is also a feature of the node.
    -> This function takes an existing network, with whatever nodes it has
       and adds a senator to it.
    -> You can create a network with other nodes and the senate to it.
    -> This function currently adds no connections to the nodes.
    -> Name of features:
         'ideology'
         'state'

    Parameters
    ----------
    network: Cayley network object
       A graph object (prefered) or lattice or cayley tree.

    csv_name: str default='senatedata.csv'
       CSV file that stores the data of the senate. A copy can be obtained at the
       following link: 
       https://www.govtrack.us/congress/members/report-cards/2017/senate/ideology
       It must be stored in same folder as this code.
       
    Examples
    --------
    >>> import Cayley as cy
    >>> import Cayley.game_theory as cgt
    >>> g = cy.Graph()
    >>> cgt.senate(g)
    """
    with open(csv_name) as csvfile:
        readCSV = csv.reader(csvfile, delimiter=',')
        for row in readCSV:
            if row[8] != 'name':
                network.add(row[8],ideology = row[3],state = row[6])

def payoff_matrix1(network,a,b,c,d,agent1,agent2):
    """
    Calculates the payoff matrix when the issue value is less than 0.5 and
    the main agent playing the game has an idealogical score of less than 0.5.

    Parameters
    ----------
    network: Cayley network object
       A graph object (prefered) or lattice or cayley tree.
    a: float
       This is a constant that goes into calculating the values in the matrix.
    b: float
       This is a constant that goes into calculating the values in the matrix.
    c: float
       This is a constant that goes into calculating the values in the matrix.
    d: float
       This is a constant that goes into calculating the values in the matrix.
    agent1: str
       A string that contains the exact last name of the senator that appears
       in the network must be used. Proper capitalization must be used.
    agent2: str
       A string that contains the exact last name of the senator that appears
       in the network must be used. Proper capitalization must be used.
       
    Notes
    -----
    -> The agent listed first will be the vertical strategy. If this is not correct,
       the results will not be as intended!
    -> Agent listed second will be the horizontal strategy.
    -> The numpy package needs to be installed in order for function to work.
    -> The matrix returned is a numpy matrix object. This allows us to use their
       interface to do many linear algebra operations as well as array operations. 
    -> Conditions of issue rating and agent idealogial score must be checked
       before the funciton is executed.

    Returns
    -------
    A numpy matrix

    Raises
    ------
    KeyError:
       When the ideology feature was not added to the entire network.

    Examples
    --------
    >>> import Cayley as cy
    >>> import Cayley.game_theory as cgt
    >>> g = cy.Graph()
    >>> cgt.senate(g)
    >>> cgt.payoff_matrix1(0.5,1.2,3.3,1.0,"Cruz","Sanders")
    """
    try:
        a1 = float(network.getNodeFeature('ideology')[agent1])
        a2 = float(network.getNodeFeature('ideology')[agent2])
        matrix1 = np.matrix([[a*abs(a1-a2),b*abs(a1-a2)],
                             [c*abs(a1-a2),d*abs(a1-a2)]],dtype = float)
        return matrix1
    except KeyError:
        return "Ideology feature not added to the entire network."
    
def payoff_matrix2(network,a,b,c,d,agent1,agent2):
    """
    Calculates the payoff matrix when the issue value is less than 0.5 and
    the main agent playing the game has an idealogical score of greater than 0.5.

    Parameters
    ----------
    network: Cayley network object
       A graph object (prefered) or lattice or cayley tree.
    a: float
       This is a constant that goes into calculating the values in the matrix.
    b: float
       This is a constant that goes into calculating the values in the matrix.
    c: float
       This is a constant that goes into calculating the values in the matrix.
    d: float
       This is a constant that goes into calculating the values in the matrix.
    agent1: str
       A string that contains the exact last name of the senator that appears
       in the network must be used. Proper capitalization must be used.
    agent2: str
       A string that contains the exact last name of the senator that appears
       in the network muyst be used. Proper capitalization must be used.
       
    Notes
    -----
    -> The agent listed first will be the vertical strategy. If this is not correct,
       the results will not be correct!
    -> Agent listed second will be the horizontal strategy.
    -> The numpy package needs to be installed in order for function to work.
    -> The matrix returned is a numpy matrix object. This allows us to use their
       interface to do many linear algebra operations as well as array operations.
    -> Conditions of issue rating and agent idealogial score must be checked
       before the funciton is executed.

    Returns
    -------
    A numpy matrix

    Raises
    ------
    KeyError:
       When the ideology feature was not added to the entire network.

    Examples
    --------
    >>> import Cayley as cy
    >>> import Cayley.game_theory as cgt
    >>> g = cy.Graph()
    >>> cgt.senate(g)
    >>> cgt.payoff_matrix2(0.5,1.2,3.3,1.0,"Cruz","Sanders")
    """
    try:
        a1 = float(network.getNodeFeature('ideology')[agent1])
        a2 = float(network.getNodeFeature('ideology')[agent2])
        matrix2 = np.matrix([[d*abs(a1-a2),c*abs(a1-a2)],
                             [b*abs(a1-a2),a*abs(a1-a2)]],dtype = float)
        return matrix2
    except KeyError:
        return "Ideology feature not added to the entire network."
    
def payoff_matrix3(network,a,b,c,d,agent1,agent2):
    """
    Calculates the payoff matrix when the issue value is greater than 0.5 and
    the main agent playing the game has an idealogical score of less than 0.5.

    Parameters
    ----------
    network: Cayley network object
       A graph object (prefered) or lattice or cayley tree.
    a: float
       This is a constant that goes into calculating the values in the matrix.
    b: float
       This is a constant that goes into calculating the values in the matrix.
    c: float
       This is a constant that goes into calculating the values in the matrix.
    d: float
       This is a constant that goes into calculating the values in the matrix.
    agent1: str
       A string that contains the exact last name of the senator that appears
       in the network must be used. Proper capitalization must be used.
    agent2: str
       A string that contains the exact last name of the senator that appears
       in the network must be used. Proper capitalization must be used.
       
    Notes
    -----
    -> The agent listed first will be the vertical strategy. If this is not correct,
       the results will not be correct!
    -> Agent listed second will be the horizontal strategy.
    -> The numpy package needs to be installed in order for function to work.
    -> The matrix returned is a numpy matrix object. This allows us to use their
       interface to do many linear algebra operations as well as array operations. 
    -> Conditions of issue rating and agent idealogial score must be checked
       before the funciton is executed.

    Returns
    -------
    A numpy matrix

    Raises
    ------
    KeyError:
       When the ideology feature was not added to the entire network.

    Examples
    --------
    >>> import Cayley as cy
    >>> import Cayley.game_theory as cgt
    >>> g = cy.Graph()
    >>> cgt.senate(g)
    >>> cgt.payoff_matrix3(0.5,1.2,3.3,1.0,"Cruz","Sanders")
    """
    try:
        a1 = float(network.getNodeFeature('ideology')[agent1])
        a2 = float(network.getNodeFeature('ideology')[agent2])
        matrix3 = np.matrix([[d*abs(a1-a2),c*abs(a1-a2)],
                             [b*abs(a1-a2),a*abs(a1-a2)]],dtype = float)
        return matrix3
    except KeyError:
        return "Ideology feature not added to the entire network."

def payoff_matrix4(network,a,b,c,d,agent1,agent2):
    """
    Calculates the payoff matrix when the issue value is greater than 0.5 and
    the main agent playing the game has an idealogical score of greater than 0.5.

    Parameters
    ----------
    network: Cayley network object
       A graph object (prefered) or lattice or cayley tree.
    a: float
       This is a constant that goes into calculating the values in the matrix.
    b: float
       This is a constant that goes into calculating the values in the matrix.
    c: float
       This is a constant that goes into calculating the values in the matrix.
    d: float
       This is a constant that goes into calculating the values in the matrix.
    agent1: str
       A string that contains the exact last name of the senator that appears
       in the network must be used. Proper capitalization must be used.
    agent2: str
       A string that contains the exact last name of the senator that appears
       in the network must be used. Proper capitalization must be used.
       
    Notes
    -----
    -> The agent listed first will be the vertical strategy. If this is not correct,
       the results will not be correct!
    -> Agent listed second will be the horizontal strategy.
    -> The numpy package needs to be installed in order for function to work.
    -> The matrix returned is a numpy matrix object. This allows us to use their
       interface to do many linear algebra operations as well as array operations.
    -> Conditions of issue rating and agent idealogial score must be checked
       before the funciton is executed.

    Returns
    -------
    A numpy matrix

    Raises
    ------
    KeyError:
       When the ideology feature was not added to the entire network.

    Examples
    --------
    >>> import Cayley as cy
    >>> import Cayley.game_theory as cgt
    >>> g = cy.Graph()
    >>> cgt.senate(g)
    >>> cgt.payoff_matrix4(0.5,1.2,3.3,1.0,"Cruz","Sanders")
    """
    try:
        a1 = float(network.getNodeFeature('ideology')[agent1])
        a2 = float(network.getNodeFeature('ideology')[agent2])
        matrix4 = np.matrix([[a*abs(a1-a2),b*abs(a1-a2)],
                             [c*abs(a1-a2),d*abs(a1-a2)]],dtype = float)
        return matrix4
    except KeyError:
        return "Ideology feature not added to the entire network."

def random_strat_start(network):
    """
    Randomly chooses the strategy of the agent. A zero will represent no
    and a one will represent yes. Will add the strategy as a feature to
    each senator node.

    Notes
    -----
    -> The name of the feature is 'strategy'

    Parameters
    ----------
    network: Cayley network object
       A graph object (prefered) or lattice or cayley tree.

    Examples
    --------
    >>> import Cayley as cy
    >>> import Cayley.game_theory as cgt
    >>> g = cy.Graph()
    >>> senate(g)
    >>> cgt.randomStart(g)
    """
    for node in network:
        network.add(node, strategy = random.randint(0,1))

def game(network,payoff_matrix,agent1,agent2,strategy_d):
    """
    Plays a game between two agents. Agent 1 is playing the vertical strategy
    and agent 2 is playing the horizontal strategy. The function adds the
    real reward and the imagined reward as features to the senator node that
    is agent 1.

    Parameters
    ----------
    network: Cayley network object
       A graph object (prefered) or lattice or cayley tree.
    payoff_matrix: numpy matrix object
       A numpy matrix must be used otherwise the code will not run!
    agent1: str
       Name of agent 1 as it appears in the network.
    agent2: str
       Name of agent 2 as it appears in the network.

    Returns
    -------
    Tuple with the real reward of agent 1 as the first element and the imagined
    reward as the second element. Both are floats.

    Raises
    ------
    KeyError:
       Raises when a strategy feature has not been assigned to the entire network.

    Notes
    -----
    -> The payoff matrix must be determined before with the conditions of the
       issue value and agent 1 ideological value. But it will take any matrix.
    -> The agent listed first will be the vertical strategy. If this is not correct,
       the results will not be correct!
    -> Agent listed second will be the horizontal strategy.
    -> The agents listed in the function arguemnts do effect calculations,
       since the real and imagined rewards are only added to the agent1 argument.
    -> The real reward and imagined rewards are added as features to the senator
       nodes, listed as agent1, under the names:
            'real_reward'
            'imagined_reward'

    Examples
    --------
    >>> import Cayley as cy
    >>> import Cayley.game_theory as cgt
    >>> g = cy.Graph()
    >>> cgt.senate(g)
    >>> matrix = cgt.payoff_matrix4(0.5,1.2,3.3,1.0,"Cruz","Sanders")
    >>> game(matrix,"Cruz","Sanders")
    """
    try:
        agent1_strategy = strategy_d[agent1]
        agent2_strategy = strategy_d[agent2]
        real_payoff = payoff_matrix[1-agent2_strategy,1-agent1_strategy]
        imagined_payoff = payoff_matrix[1-agent2_strategy,agent1_strategy]
        network.add(agent1,imagined_reward = imagined_payoff,
                    real_reward = real_payoff)
    except KeyError:
        return "Strategy feature is not built into entire network"

def random_agent(network,agent1):
    """
    Chooses a random agent relative to another agent. Meaning, if one
    picks 'A' and wants another random agent to make a pair, 'A' will not
    be chosen. Sets up a connection between two random agents.

    Parameters
    ----------
    network: Cayley network object
       A graph object (prefered) or lattice or cayley tree.
    agent1: str
       Name of agent 1 as it appears in the network.

    Returns
    -------
    A name of a node in the network. In the case of game theory, another agent.

    Examples
    --------
    >>> import Cayley as cy
    >>> import Cayley.game_theory as cgt
    >>> g = cy.Graph()
    >>> cgt.senate(g)
    >>> cgt.random_agent(g,"Cruz")
    """
    agent2 = random.choice(list(network))#chooses random senator
    if agent1 == agent2:
        random_agent(network,agent1)
    network.directedLink(agent1,agent2)

def timestep(network,issue_rating,a,b,c,d,k):
    """
    Simulates a timestep of a game on an entire network. Currently it plays it
    goes through the network selects a random agent and plays a game between
    them.

    Parameters
    ----------
    network: Cayley network object
       A graph object (prefered) or lattice or cayley tree.
    issue_rating: float
       The issue rating based on the 0 to 1 scale defined in
       https://www.govtrack.us/congress/members/report-cards/2017/senate/ideology
       where the number indicates how liberal or conservative an issue is.
    a: float
       This is a constant that goes into calculating the values in payoff matrix.
    b: float
       This is a constant that goes into calculating the values in payoff matrix.
    c: float
       This is a constant that goes into calculating the values in payoff matrix.
    d: float
       This is a constant that goes into calculating the values in payoff matrix.
    k: float
       A constant used in calculting the switching probabilty as discussed in
       the equation in J. Liu et al 2017 EPL119 68001.

    Notes
    -----
    -> For every new timestep run the data from the previous is lost.
    -> Picks a random agent currently for an agent to play with, will
       change in the future.

    Examples
    --------
    >>> import Cayley as cy
    >>> import Cayley.game_theory as cgt
    >>> g = cy.Graph()
    >>> senate(g)
    >>> cgt.randomStart(g)
    >>> cgt.timestep(g,0.4,1,5.23,100,0.3,20.2)
    """
    strategy_d = network.getNodeFeature('strategy').copy()
    ideology_d = network.getNodeFeature('ideology').copy()
    for node in network:
        neighbors = network.getNodeFeature('neighbors')[node]
        #print(neighbors)
        ideology_agent1 = float(ideology_d[node])
        agent1_strategy = strategy_d[node]
        real_payoff = 0
        imagined_payoff = 0
        for senator in neighbors:
            #print(senator)
            if issue_rating < 0.5 and ideology_agent1 < 0.5:
                matrix = payoff_matrix1(network,a,b,c,d,node,senator)
            elif issue_rating < 0.5 and ideology_agent1 > 0.5:
                matrix = payoff_matrix2(network,a,b,c,d,node,senator)
            elif issue_rating > 0.5 and ideology_agent1 < 0.5:
                matrix = payoff_matrix3(network,a,b,c,d,node,senator)
            elif issue_rating > 0.5 and ideology_agent1 > 0.5:
                marix = payoff_matrix4(network,a,b,c,d,node,senator)
            agent2_strategy = strategy_d[senator]
            real_payoff += matrix[1-agent2_strategy,1-agent1_strategy]
            imagined_payoff += matrix[1-agent2_strategy,agent1_strategy]
        network.add(node,imagined_reward = imagined_payoff,
                    real_reward = real_payoff)
        #game(network,matrix,node,agent2,strategy_d)
        #using strat_d as arg ensures copy is used
        imagined_d = network.getNodeFeature('imagined_reward').copy()
        real_d = network.getNodeFeature('real_reward').copy()
        #print("senator: ",node)
        #print("Imagined: ",imagined_d[node])
        #print("Real: ",real_d[node])
        if imagined_d[node] > real_d[node]:
            network.add(node,strategy = 1-strategy_d[node])
        else:
            switch_probability = math.e**(-1*(real_d[node]-imagined_d[node])/k)
            if random.uniform(0,1) <= switch_probability:
                network.add(node,strategy = 1-strategy_d[node])
    #print(network.getNodeFeature('strategy'))

def data_export(name,network):
    """
    Saves that current lies within the strategy, imagined reward, and real
    rewards, assuming they exist. If spreadsheet with same name exists, it
    will add the data to the already existing spreadsheet.

    Parameters
    ----------
    name: str
       The name of the excel spreadsheet
    network: Cayley network object
       A graph object (prefered) or lattice or cayley tree.

    Notes
    -----
    -> Best to export data to an excel sheet after running a timestep
    -> Inital strategies for agents are not exported to the spreadsheet.    
    """
    strategy_d = network.getNodeFeature('strategy')
    imagined_d = network.getNodeFeature('imagined_reward')
    real_d = network.getNodeFeature('real_reward')
    try: #checks to see if excel sheet exists to add to existing simulation
        book = openpyxl.load_workbook(name+'.xlsx')
        sheet = book['Strategy']
        col = sheet.max_column
        strat_data(strategy_d,col,book,network)
        real_reward_data(real_d,col,book,network)
        imagined_reward_data(imagined_d,col,book,network)
        book.save(name+'.xlsx')
    except FileNotFoundError:
        strat_real_imagined_setup(name,network)
        #adds first set of data from end of 0 timestep
        book = openpyxl.load_workbook(name+'.xlsx')
        strat_data(strategy_d,1,book,network)
        real_reward_data(real_d,1,book,network)
        imagined_reward_data(imagined_d,1,book,network)
        book.save(name+'.xlsx')

def strat_real_imagined_setup(name,network):
    """
    If other types of spreadsheets need to be created where data from
    strategy, real reward and imagined reward need to be used, this creates
    the initial structure.

    Parameters
    ----------
    name: str
       The name of the excel spreadsheet
    network: Cayley network object
       A graph object (prefered) or lattice or cayley tree.

    Notes
    -----
    -> This works as a standalone function and can be built on top of. 
    """
    book = openpyxl.Workbook()
    std=book['Sheet']
    book.remove(std)
    #creates strategy sheet
    book.create_sheet('Strategy')
    sheet = book['Strategy']
    sheet.cell(row=1, column=1).value = 'Timestep'
    #creates real rewards sheet
    book.create_sheet('Real Rewards')
    real_sheet = book['Real Rewards']
    real_sheet.cell(row=1, column=1).value = 'Timestep' 
    #creates imagined rewards sheet
    book.create_sheet('Imagined Rewards')
    imagined_sheet = book['Imagined Rewards']
    imagined_sheet.cell(row=1, column=1).value = 'Timestep'
    #adds senators
    row_count = 2
    for senator in network:
        sheet.cell(row=row_count,column=1).value = senator
        real_sheet.cell(row=row_count,column=1).value = senator
        imagined_sheet.cell(row=row_count,column=1).value = senator
        row_count += 1
    #saves excel sheet
    book.save(name+'.xlsx')

def strat_data(data,col,book,network):
    """
    This function is a not a standalone function. An excel spreadsheet must
    be created with the appropriate sheet name. This exists to make the code
    in other functions more readable since there was a lot of repeat. 
    """
    sheet = book['Strategy']
    sheet.cell(row=1, column=col+1).value = col-1
    row_count = 2
    for senator in network:
        sheet.cell(row=row_count,column=col+1).value = data[senator]
        row_count+=1

def real_reward_data(data,col,book,network):
    """
    This function is a not a standalone function. An excel spreadsheet must
    be created with the appropriate sheet name. This exists to make the code
    in other functions more readable since there was a lot of repeat. 
    """
    real_sheet = book['Real Rewards']
    real_sheet.cell(row=1, column=col+1).value = col-1
    row_count = 2
    for senator in network:
        real_sheet.cell(row=row_count,column=col+1).value = data[senator]
        row_count+=1

def imagined_reward_data(data,col,book,network):
    """
    This function is a not a standalone function. An excel spreadsheet must
    be created with the appropriate sheet name. This exists to make the code
    in other functions more readable since there was a lot of repeat. 
    """
    imagined_sheet = book['Imagined Rewards']
    imagined_sheet.cell(row=1, column=col+1).value = col-1
    row_count = 2
    for senator in network:
        imagined_sheet.cell(row=row_count,column=col+1).value = data[senator]
        row_count+=1

def export_data(name,network,data1,data2,data3):
    """
    Exports data to excel. Takes the network representing the senate
    as a parameter as well as the data set representing the strategy, the
    data set representing the real rewards, and the data set representing the
    imagined rewards. 
    """
    workbook = xlsxwriter.Workbook(name+'.xlsx')
    worksheet = workbook.add_worksheet("Strategy")
    worksheet.write(0,0,"Timestep")
    row_count = 1
    for senator in network:
        worksheet.write(row_count,0,senator)
        row_count += 1
    column_count = 1
    for y in range(len(data1)):
        worksheet.write(0,column_count,str(y))
        column_count += 1
    for y in range(len(data1)):
        row_count2 = 1
        for senator in network:
            worksheet.write(row_count2,y+1,data1[y][senator])
            row_count2 += 1
            
    worksheet2 = workbook.add_worksheet("Real Rewards")
    worksheet2.write(0,0,"Timestep")
    row_count3 = 1
    for senator in network:
        worksheet2.write(row_count3,0,senator)
        row_count3 += 1
    column_count = 1
    for y in range(len(data2)):
        worksheet2.write(0,column_count,str(y))
        column_count += 1
    for y in range(len(data2)):
        row_count4 = 1
        for senator in network:
            worksheet2.write(row_count4,y+1,data2[y][senator])
            row_count4 += 1
    
    worksheet3 = workbook.add_worksheet("Imagined Rewards")
    worksheet3.write(0,0,"Timestep")
    row_count5 = 1
    for senator in network:
        worksheet3.write(row_count5,0,senator)
        row_count5 += 1
    column_count = 1
    for y in range(len(data3)):
        worksheet3.write(0,column_count,str(y))
        column_count += 1
    for y in range(len(data3)):
        row_count6 = 1
        for senator in network:
            worksheet3.write(row_count6,y+1,data3[y][senator])
            row_count6 += 1            
    workbook.close()
    
    
